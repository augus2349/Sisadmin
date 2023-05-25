from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import Label
from tkinter import Entry
import tkinter.font as tkFont
from openpyxl import *
from openpyxl import load_workbook

horas=["07:00","07:15","07:30","07:45","08:00","08:15","08:30","08:45","09:00",
       "09:15","09:30","09:45","10:00","10:15","10:30","10:45","11:00","11:15",
       "11:30","11:45","12:00","12:15","12:30","12:45","13:00","13:15","13:30",
       "13:45","14:00","14:15","14:30","14:45","15:00","15:15","15:30","15:45",
       "16:00","16:15","16:30","16:45","17:00","17:15","17:30","17:45","18:00",
       "18:15","18:30","18:45","19:00","19:15","19:30","19:45","20:00","20:15",
       "20:30","20:45","21:00"]
mes=["01","02","03","04","05","06","07","08","09","10","11","12","XX"]
ano=["22","23","24","25","26","27","28","29","30","31","32","33","34","35",
     "36","37","38","39","40","XX"]
dia=["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15",
     "16","17","18","19","20","21","22","23","24","25","26","27","28","29","30",
     "31"]
wb1 = load_workbook('Empleado.xlsx',data_only=True)
ws1= wb1['Puesto']
ws2= wb1['Horario']
ws3= wb1['Vacaciones']
ws4= wb1['Facturas']
wb2 = load_workbook('Almacen.xlsx',data_only=True)
ht1= wb2['Disponibles']
ht2= wb2['Uso']
ht3= wb2['Pedidos']
ht4=wb2['Facturas']

home = Tk()
home.title("Sistema Administrativo")
home.resizable(0,0)
fuente=tkFont.Font(family="Bahnschrift SemiLight")

def rec():
    def agre():
        c=0
        i=2
        w=0
        if numr.get()=="" or nomr.get()=="" or aper.get()=="" or puer.get()=="" or arer.get()=="" or supr.get()=="" or dtr.get()=="":
            c=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para completar el registro")
            w=1
        while c==0:   
            a=ws1[f'A{i}']
            v=a.value
            print(v)
            if v==None:
                c=1
                ws1[f'A{i}']=numr.get()
                ws2[f'A{i}']=numr.get()
                ws3[f'A{i}']=numr.get()
                ws1[f'B{i}']=nomr.get()
                ws2[f'B{i}']=nomr.get()
                ws3[f'B{i}']=nomr.get()
                ws1[f'C{i}']=aper.get()
                ws2[f'C{i}']=aper.get()
                ws3[f'C{i}']=aper.get()
                ws1[f'D{i}']=puer.get()
                ws2[f'D{i}']=puer.get()
                ws1[f'F{i}']=arer.get()
                ws1[f'G{i}']=supr.get()
                ws2[f'E{i}']=hinr.get()
                ws2[f'F{i}']=houtr.get()
                ws2[f'G{i}']=hcomr.get()
                ws2[f'H{i}']=dtr.get()
            elif v==numr.get():
                messagebox.showinfo("ADVERTENCIA","YA EXISTE UN REGISTRO\nCON ESE NUMERO DE EMPLEADO")
                c=1
            else:
                i=i+1
        if w==0:
            numr.set("")
            nomr.set("")
            aper.set("")
            puer.set("")
            arer.set("")
            supr.set("")
            dtr.set("")
        wb1.save('Empleado.xlsx')
    
    def consul():
        q=0
        i=2
        if numr.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","No se ingresó algún número para consultar")
        while q==0:   
            a=ws1[f'A{i}']
            v=a.value
            print(v)
            b=ws1[f'B{i}']
            c=ws1[f'C{i}']
            d=ws1[f'D{i}']
            e=ws1[f'F{i}']
            f=ws1[f'G{i}']
            g=ws2[f'E{i}']
            h=ws2[f'F{i}']
            j=ws2[f'G{i}']
            k=ws2[f'H{i}']

            if v==numr.get():
                nomr.set(b.value)
                aper.set(c.value)
                puer.set(d.value)
                arer.set(e.value)
                supr.set(f.value)
                hinr.set(g.value)
                houtr.set(h.value)
                hcomr.set(j.value)
                dtr.set(k.value)
                q=1
            elif v==None:
                messagebox.showinfo("ADVERTENCIA","No hay un empleado asignado a este numero")
                q=1
            else:
                i=i+1 
                    
    emp=Toplevel()
    emp.title("Registro de Asistencia")
    #emp.geometry("950x360")
    note= ttk.Notebook(emp)
    note.pack(fill='both',expand='yes')
    
#-------------------------registro de nuevo usuario----------------------------
    numr=StringVar()
    nomr=StringVar()
    aper=StringVar()
    puer=StringVar()
    arer=StringVar()
    supr=StringVar()
    hinr=StringVar()
    hinr.set(horas[0])
    houtr=StringVar()
    houtr.set(horas[32])
    hcomr=StringVar()
    hcomr.set(horas[17])
    dtr=StringVar()
    rne=Frame(note,bg="gray")
    nume=Label(rne, font="Fuente 15", text="N° Empleado:",bg="gray")
    nume.grid(row=1,column=0,padx=10,pady=15)
    num=Entry(rne,font="Fuente 15", width="10",textvariable=numr)
    num.grid(row=1,column=1,sticky=W)
    nomb=Label(rne, font="Fuente 15", text="Nombre:",bg="gray")
    nomb.grid(row=2,column=0,padx=10)
    nom=Entry(rne,font="Fuente 15", width="30",textvariable=nomr)
    nom.grid(row=2,column=1,columnspan=4)
    apel=Label(rne, font="Fuente 15", text="Apellido:",bg="gray")
    apel.grid(row=3,column=0,padx=10,pady=15)
    ape=Entry(rne,font="Fuente 15", width="30",textvariable=aper)
    ape.grid(row=3,column=1,columnspan=4)
    pues=Label(rne, font="Fuente 15", text="Puesto:",bg="gray")
    pues.grid(row=4,column=0,padx=10)
    pue=Entry(rne,font="Fuente 15", width="30",textvariable=puer)
    pue.grid(row=4,column=1,columnspan=4)
    area=Label(rne, font="Fuente 15", text="Área:",bg="gray")
    area.grid(row=5,column=0,padx=10,pady=15)
    are=Entry(rne,font="Fuente 15", width="30",textvariable=arer)
    are.grid(row=5,column=1,columnspan=4)
    supe=Label(rne, font="Fuente 15", text="Superior:",bg="gray")
    supe.grid(row=6,column=0,padx=10)
    sup=Entry(rne,font="Fuente 15", width="30",textvariable=supr)
    sup.grid(row=6,column=1,columnspan=4)
    h=Label(rne, font="Fuente 15", text="        Horario",bg="gray")
    h.grid(row=1,column=5,columnspan=4)
    h_in=Label(rne, font="Fuente 15", text="   Entrada   ",bg="gray")
    h_in.grid(row=2,column=5,padx=5)
    hin=OptionMenu(rne,hinr,*horas)
    hin.config(font="Fuente 15", width="5")
    hin.grid(row=2,column=6)
    h_out=Label(rne, font="Fuente 15", text="Salida   ",bg="gray")
    h_out.grid(row=2,column=7,padx=5)
    hout=OptionMenu(rne,houtr,*horas)
    hout.config(font="Fuente 15", width="5")
    hout.grid(row=2,column=8)
    h_com=Label(rne, font="Fuente 15", text="   Comida   ",bg="gray")
    h_com.grid(row=3,column=5)
    hcom=OptionMenu(rne,hcomr,*horas)
    hcom.config(font="Fuente 15", width="5")
    hcom.grid(row=3,column=6)
    dt=Label(rne, font="Fuente 15", text="Días\n   Laborales   ",bg="gray")
    dt.grid(row=3,column=7)
    dts=Entry(rne,font="Fuente 15", width="8",textvariable=dtr)
    dts.grid(row=3,column=8)
    rn=Button(rne,command=lambda:agre())
    rn.config(font="fuente 15",text="Registrar\nNuevo Empleado",height=2)
    rn.grid(row=5,column=5,rowspan=2,columnspan=2)
    cd=Button(rne,command=lambda:consul())
    cd.config(font="fuente 15",text="Consultar Datos\ndel Empleado",height=2)
    cd.grid(row=5,column=7,rowspan=2,columnspan=2)
    v=Label(rne, font="Fuente 15", text="     ",bg="gray")
    v.grid(row=7,column=0)

#-------------------------------------Registro de asistencia---------------------
    def rasi():
        q=0
        i=2
        if numas.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","No se ingresó algún número para registrar")
        while q==0:   
            a=ws3[f'A{i}']
            v=a.value
            if v==numas.get():
                tasi=ws3[f'D{i}']
                if tasi.value==None:
                    asi=0
                else:
                    asi=int(tasi.value)
                nasi=asi+1
                q=1
                ws3[f'D{i}']=nasi
                messagebox.showinfo("Tarea exitosa","Asistencia Registrada")
            elif v==None:
                messagebox.showinfo("ADVERTENCIA","No hay un empleado asignado a este numero")
                q=1
            else:
                i=i+1 
        
        wb1.save('Empleado.xlsx')
        numas.set("")

    numas=StringVar()    
    hinas=StringVar()
    hinas.set(horas[0])
    houtas=StringVar()
    houtas.set(horas[32])
    ras=Frame(note,bg="gray")
    nume=Label(ras, font="Fuente 15", text="N° Empleado:",bg="gray")
    nume.grid(row=1,column=1,pady=15,columnspan=2,sticky=E)
    num=Entry(ras,font="Fuente 15", width="10",textvariable=numas)
    num.grid(row=1,column=3,padx=15,columnspan=2,sticky=W)
    h_ins=Label(ras,font="Fuente 15", text="Hora de\nEntrada:",bg="gray")
    h_ins.grid(row=2,column=1,padx=10,pady=15)
    hins=OptionMenu(ras,hinas,*horas)
    hins.config(font="Fuente 15", width="5")
    hins.grid(row=2,column=2,padx=10)
    h_out=Label(ras,font="Fuente 15", text="Hora de\nSalida:",bg="gray")
    h_out.grid(row=2,column=3,padx=10,pady=15)
    houts=OptionMenu(ras,houtr,*horas)
    houts.config(font="Fuente 15", width="5")
    houts.grid(row=2,column=4,padx=10)
    rege=Button(ras,command=lambda:rasi())
    rege.config(font="fuente 15",text="Registrar\nAsistencia", width="15",height=2)
    rege.grid(row=3,column=2,columnspan=2,padx=10)

#------------------------gestion vacacional------------------------------------
    
    gesv=Frame(note,bg="gray")
    
    def solic():
        q=0
        i=2
        if nuvac.get()=="" or numvac.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","No se ingresó algún número para registrar la solicitud")
        while q==0:   
            a=ws3[f'A{i}']
            v=a.value
            b=ws3[f'B{i}']
            c=ws3[f'C{i}']
            d=ws3[f'D{i}']
            e=numvac.get()
            f="Pendiente"
            if v==nuvac.get():
                if (ws3[f'E{i}']).value==None:
                    ws3[f'E{i}']=e
                    ws3[f'F{i}']=f
                    lvac.insert("",END,text=v,values=(b.value,c.value,d.value,e,f))
                    q=1
                else:
                    messagebox.showinfo("ADVERTENCIA","Ya existe una solicitud")
                    q=1
            elif v==None:
                messagebox.showinfo("ADVERTENCIA","No hay un empleado asignado a este numero")
                q=1
            else:
                i=i+1
            wb1.save('Empleado.xlsx')
    def aprov():
        selected = lvac.focus()
        temp = lvac.item(selected, 'values')
        lvac.item(selected, values=(temp[0], temp[1], temp[2],temp[3],"Aprovada"))
        valin=lvac.item(selected, 'text')
        q=0
        i=2
        while q==0:   
            a=ws3[f'A{i}']
            v=a.value
            if v==valin:
                ws3[f'F{i}']="Aprovada"
                q=1
            else:
                i=i+1
            wb1.save('Empleado.xlsx')
    def recha():
        selected = lvac.focus()
        temp = lvac.item(selected, 'values')
        lvac.item(selected, values=(temp[0], temp[1], temp[2],temp[3],"Rechazar"))
        valin=lvac.item(selected, 'text')
        q=0
        i=2
        while q==0:   
            a=ws3[f'A{i}']
            v=a.value
            if v==valin:
                ws3[f'F{i}']="Rechazar"
                q=1
            else:
                i=i+1
            wb1.save('Empleado.xlsx')
                
    nuvac=StringVar()
    numvac=StringVar()
    nvac=Label(gesv,font="Fuente 15", text="N° Empleado:",bg="gray")
    nvac.grid(row=1,column=0,pady=15,padx=10)
    n_vac=Entry(gesv,font="Fuente 15", width="10",textvariable=nuvac)
    n_vac.grid(row=1,column=1,sticky=W)
    dsol=Label(gesv,font="Fuente 15", text="Dias a Solicitar:",bg="gray")
    dsol.grid(row=1,column=2,pady=15,padx=10)
    d_sol=Entry(gesv,font="Fuente 15", width="10",textvariable=numvac)
    d_sol.grid(row=1,column=3,sticky=W)
    soli=Button(gesv,command=lambda:solic())
    soli.config(font="fuente 15",text="Solicitar", width="15")
    soli.grid(row=1,column=4,padx=30,pady=15)
    lvac=ttk.Treeview(gesv,columns=("col1","col2","col3","col4","col5"))
    lvac.grid(row=2,column=0,columnspan=4,padx=15,pady=15,rowspan=7)
    lvac.column("#0",width="50")
    lvac.column("col1",width="100")
    lvac.column("col2",width="100")
    lvac.column("col3",width="80")
    lvac.column("col4",width="80")
    lvac.column("col5",width="100")
    lvac.heading("#0",text="N° Emp",anchor=CENTER)
    lvac.heading("col1",text="Nombre",anchor=CENTER)
    lvac.heading("col2",text="Apellido",anchor=CENTER)
    lvac.heading("col3",text="Dias Acumulados",anchor=CENTER)
    lvac.heading("col4",text="Dias Solicitados",anchor=CENTER)
    lvac.heading("col5",text="Estatus",anchor=CENTER)
    apro=Button(gesv,command=lambda:aprov())
    apro.config(font="fuente 15",text="Aprovar", width="15")
    apro.grid(row=3,column=4,padx=30,pady=15)
    rech=Button(gesv,command=lambda:recha())
    rech.config(font="fuente 15",text="Rechazar", width="15")
    rech.grid(row=4,column=4,padx=30,pady=15)
    c=0
    i=2
    while c==0:
        Dt=(ws3[f'A{i}']).value
        Et=(ws3[f'B{i}']).value
        Ft=(ws3[f'C{i}']).value
        Gt=(ws3[f'D{i}']).value
        It=(ws3[f'E{i}']).value
        Jt=(ws3[f'F{i}']).value

            
        a=ws3[f'E{i}']
        v=a.value
        if v==None:
            c=1
        else:
            lvac.insert("",END,text=Dt,values=(Et,Ft,Gt,It,Jt))
            i=i+1
    
    
#--------------------------Cambio de Puesto------------------------------------
    camp=Frame(note,bg="gray")
    
    def campu():
        q=0
        i=2
        if numc.get()=="" or puec.get()=="" or arec.get()=="" or supc.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","Faltan elementos para el cambio")
        while q==0:   
            a=ws1[f'A{i}']
            v=a.value
            if v==numc.get():
                pant=(ws1[f'D{i}'].value)
                ws1[f'D{i}']=puec.get()
                ws1[f'E{i}']=pant
                ws1[f'F{i}']=arec.get()
                ws1[f'G{i}']=supc.get()
                q=1
            elif v==None:
                messagebox.showinfo("ADVERTENCIA","No hay un empleado asignado a este numero")
                q=1
            else:
                i=i+1 
        wb1.save('Empleado.xlsx')
        puec.set("")
        numc.set("")
        arec.set("")
        supc.set("")
        
    numc=StringVar()
    puec=StringVar()
    arec=StringVar()
    supc=StringVar()
    nume=Label(camp, font="Fuente 15", text="N° Empleado:",bg="gray")
    nume.grid(row=1,column=0,pady=15,padx=10)
    num=Entry(camp,font="Fuente 15", width="10",textvariable=numc)
    num.grid(row=1,column=1,sticky=W)
    pues=Label(camp, font="Fuente 15", text="Puesto:",bg="gray")
    pues.grid(row=4,column=0,padx=10)
    pue=Entry(camp,font="Fuente 15", width="30",textvariable=puec)
    pue.grid(row=4,column=1,columnspan=4)
    area=Label(camp, font="Fuente 15", text="Area:",bg="gray")
    area.grid(row=5,column=0,padx=10,pady=15)
    are=Entry(camp,font="Fuente 15", width="30",textvariable=arec)
    are.grid(row=5,column=1,columnspan=4)
    supe=Label(camp, font="Fuente 15", text="Superior:",bg="gray")
    supe.grid(row=6,column=0,padx=10)
    sup=Entry(camp,font="Fuente 15", width="30",textvariable=supc)
    sup.grid(row=6,column=1,columnspan=4)
    rege=Button(camp,command=lambda:campu())
    rege.config(font="fuente 15",text="Registrar\nCambio", width="15",height=2)
    rege.grid(row=7,column=1,columnspan=2,padx=10, pady=15)
    
#------------------------------Datos Globales----------------------------------
    datg=Frame(note,bg="gray")
    ldate=ttk.Treeview(datg,columns=("col1","col2","col3","col4","col5"))
    ldate.grid(row=2,column=0,columnspan=4,padx=15,pady=15)
    ldate.column("#0",width="50")
    ldate.column("col1",width="100")
    ldate.column("col2",width="100")
    ldate.column("col3",width="100")
    ldate.column("col4",width="50")
    ldate.column("col5",width="100")
    ldate.heading("#0",text="N° Emp",anchor=CENTER)
    ldate.heading("col1",text="Nombre",anchor=CENTER)
    ldate.heading("col2",text="Apellido",anchor=CENTER)
    ldate.heading("col3",text="Puesto",anchor=CENTER)
    ldate.heading("col4",text="Area",anchor=CENTER)
    ldate.heading("col5",text="Supervisor",anchor=CENTER)
    c=0
    i=2
    while c==0:
        Dt=(ws1[f'A{i}']).value
        Et=(ws1[f'B{i}']).value
        Ft=(ws1[f'C{i}']).value
        Gt=(ws1[f'D{i}']).value
        It=(ws1[f'F{i}']).value
        Jt=(ws1[f'G{i}']).value

            
        a=ws1[f'A{i}']
        v=a.value
        if v==None:
            c=1
        else:
            ldate.insert("",END,text=Dt,values=(Et,Ft,Gt,It,Jt))
            i=i+1
    
    
    
    note.add(rne,text='Registro de\n empleado')
    note.add(ras,text='Registro de\n asistencia')
    note.add(camp,text='Cambio de\n    Puesto')
    note.add(gesv,text='  Gestion\nvacacional')
    note.add(datg,text="Datos\nGlobales")

def face():
    face=Toplevel()
    face.title("Pagos aEmpleados/Facturas")
    note= ttk.Notebook(face)
    note.pack(fill='both',expand='yes')
    
#-----------------------------Factura nueva------------------------------------
    facn=Frame(note,bg="gray")
    def agref():
        c=0
        i=2
        w=0
        if numef.get()=="" or nombf.get()=="" or apelf.get()=="" or montf.get()=="" or areaf.get()=="" or encaf.get()=="" :
            c=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while c==0:   
            a=ws4[f'A{i}']
            v=a.value
            if v==None:
                ws4[f'A{i}']=numef.get()
                ws4[f'B{i}']=nombf.get()
                ws4[f'C{i}']=apelf.get()
                ws4[f'D{i}']=areaf.get()
                ws4[f'E{i}']=montf.get()
                ws4[f'F{i}']=encaf.get()
                ws4[f'G{i}']=day.get()
                ws4[f'H{i}']=month.get()
                ws4[f'I{i}']=yea.get()
                ws4[f'J{i}']=hora.get()
                c=1
            else:
                i=i+1
        if w==0:
            numef.set("")
            nombf.set("")
            apelf.set("")
            areaf.set("")
            montf.set("")
            encaf.set("")
            day.set("DD")
            month.set("MM")
            yea.set("AA")
            hora.set(horas[0])
        wb1.save('Empleado.xlsx')
    
    numef=StringVar()
    nombf=StringVar()
    apelf=StringVar()
    areaf=StringVar()
    montf=StringVar()
    encaf=StringVar()
    day=StringVar()
    month=StringVar()
    yea=StringVar()
    day.set("DD")
    month.set("MM")
    yea.set("AA")
    hora=StringVar()
    hora.set(horas[0])
    
    num=Label(facn,font="Fuente 15", text="N° Empleado:",bg="gray")
    num.grid(row=1,column=0,padx=10,pady=15)
    nume=Entry(facn,font="Fuente 15", width="10",textvariable=numef)
    nume.grid(row=1,column=1,padx=10,sticky=W)
    nomb=Label(facn,font="Fuente 15", text="Nombre:",bg="gray")
    nomb.grid(row=2,column=0,padx=10)
    nom=Entry(facn,font="Fuente 15", width="30",textvariable=nombf)
    nom.grid(row=2,column=1,columnspan=4,padx=10)
    apel=Label(facn,font="Fuente 15", text="Apellido:",bg="gray")
    apel.grid(row=3,column=0,padx=10,pady=15)
    ape=Entry(facn,font="Fuente 15", width="30",textvariable=apelf)
    ape.grid(row=3,column=1,columnspan=4,padx=10)
    area=Label(facn, font="Fuente 15", text="Area:",bg="gray")
    area.grid(row=4,column=0,padx=10)
    are=Entry(facn,font="Fuente 15", width="30",textvariable=areaf)
    are.grid(row=4,column=1,columnspan=4)
    mont=Label(facn,font="Fuente 15",text="Monto:",bg="gray")
    mont.grid(row=5,column=0,padx=10)
    mon=Entry(facn,font="Fuente 15", width="10",textvariable=montf)
    mon.grid(row=5,column=1,sticky=W,pady=15,padx=10)
    enca=Label(facn, font="Fuente 15", text="Encargado:",bg="gray")
    enca.grid(row=6,column=0,padx=10)
    enc=Entry(facn,font="Fuente 15", width="30",textvariable=encaf)
    enc.grid(row=6,column=1,columnspan=4)
    fecha=Label(facn, font="Fuente 15", text="Fecha:",bg="gray")
    fecha.grid(row=7,column=0,padx=10,pady=10)
    fdia=OptionMenu(facn,day,*dia)
    fdia.config(font="Fuente 14")
    fdia.grid(row=7,column=1)
    fmes=OptionMenu(facn,month,*mes)
    fmes.config(font="Fuente 14")
    fmes.grid(row=7,column=2,pady=10,sticky=W)
    fano=OptionMenu(facn,yea,*ano)
    fano.config(font="Fuente 14")
    fano.grid(row=7,column=3)
    h=Label(facn, font="Fuente 15", text="Hora:",bg="gray")
    h.grid(row=8,column=0,padx=10)
    hrs=OptionMenu(facn,hora,*horas)
    hrs.config(font="Fuente 14")
    hrs.grid(row=8,column=1)
    ingre=Button(facn,command=lambda:agref())
    ingre.config(font="fuente 15",text="Guardar Factura", width="15")
    ingre.grid(row=4,column=6,padx=15)
    v=Label(facn,font="Fuente 15", text="     ",bg="gray")
    v.grid(row=9,column=0)
#----------------------------Facturas mes pasado-------------------------------
    facm=Frame(note,bg="gray")
    def busc():
        c=0
        i=2
        w=0
        if femo.get()=="MM":
            c=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while c==0:
            D=(ws4[f'A{i}']).value
            E=(ws4[f'B{i}']).value
            F=(ws4[f'C{i}']).value
            G=(ws4[f'D{i}']).value
            H=(ws4[f'E{i}']).value
            I=(ws4[f'F{i}']).value
            J=(ws4[f'G{i}']).value
            K=(ws4[f'H{i}']).value
            L=(ws4[f'I{i}']).value
            M=(ws4[f'J{i}']).value
            
            a=ws4[f'H{i}']
            v=a.value
            if v==femo.get():
                lface.insert("",END,text=D,values=(E,F,G,H,I,J,K,L,M))
                i=i+1
            elif v==None:
                c=1
            else:
                i=i+1
        if w==0:
            femo.set("MM")
            
            
    femo=StringVar()
    femo.set("MM")
    fme=Label(facm,font="Fuente 15", text="Mes:",bg="gray")
    fme.grid(row=1,column=0,padx=10,pady=15)
    fmes=OptionMenu(facm,femo,*mes)
    fmes.config(font="Fuente 14")
    fmes.grid(row=1,column=1,pady=10,sticky=W)
    bus=Button(facm,command=lambda:busc())
    bus.config(font="fuente 15",text="Buscar Facturas", width="15")
    bus.grid(row=1,column=2,padx=15)
    lface=ttk.Treeview(facm,columns=("col1","col2","col3","col4","col5","col6","col7","col8","col9"))
    lface.grid(row=2,column=0,columnspan=4,padx=15)
    lface.column("#0",width="50")
    lface.column("col1",width="100")
    lface.column("col2",width="100")
    lface.column("col3",width="100")
    lface.column("col4",width="50")
    lface.column("col5",width="100")
    lface.column("col6",width="50")
    lface.column("col7",width="50")
    lface.column("col8",width="50")
    lface.column("col9",width="50")
    lface.heading("#0",text="N° Emp",anchor=CENTER)
    lface.heading("col1",text="Nombre",anchor=CENTER)
    lface.heading("col2",text="Apellido",anchor=CENTER)
    lface.heading("col3",text="Area",anchor=CENTER)
    lface.heading("col4",text="Monto",anchor=CENTER)
    lface.heading("col5",text="Encargado",anchor=CENTER)
    lface.heading("col6",text="Dia",anchor=CENTER)
    lface.heading("col7",text="Mes",anchor=CENTER)
    lface.heading("col8",text="Año",anchor=CENTER)
    lface.heading("col9",text="Hora",anchor=CENTER)
    
    
    facp=Frame(note,bg="gray")
    def carf():
        c=0
        i=2
        while c==0:
            Dt=(ws4[f'A{i}']).value
            Et=(ws4[f'B{i}']).value
            Ft=(ws4[f'C{i}']).value
            Gt=(ws4[f'D{i}']).value
            Ht=(ws4[f'E{i}']).value
            It=(ws4[f'F{i}']).value
            Jt=(ws4[f'G{i}']).value
            Kt=(ws4[f'H{i}']).value
            Lt=(ws4[f'I{i}']).value
            Mt=(ws4[f'J{i}']).value
            
            a=ws4[f'H{i}']
            v=a.value
            if v==None:
                c=1
            else:
                tface.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It,Jt,Kt,Lt,Mt))
                i=i+1
    car=Button(facp,command=lambda:carf())
    car.config(font="fuente 15",text="Buscar Facturas", width="15")
    car.grid(row=1,column=1,padx=15,pady=15)
    tface=ttk.Treeview(facp,columns=("col1","col2","col3","col4","col5","col6","col7","col8","col9"))
    tface.grid(row=2,column=0,columnspan=4,padx=15)
    tface.column("#0",width="50")
    tface.column("col1",width="100")
    tface.column("col2",width="100")
    tface.column("col3",width="100")
    tface.column("col4",width="50")
    tface.column("col5",width="100")
    tface.column("col6",width="50")
    tface.column("col7",width="50")
    tface.column("col8",width="50")
    tface.column("col9",width="50")
    tface.heading("#0",text="N° Emp",anchor=CENTER)
    tface.heading("col1",text="Nombre",anchor=CENTER)
    tface.heading("col2",text="Apellido",anchor=CENTER)
    tface.heading("col3",text="Area",anchor=CENTER)
    tface.heading("col4",text="Monto",anchor=CENTER)
    tface.heading("col5",text="Encargado",anchor=CENTER)
    tface.heading("col6",text="Dia",anchor=CENTER)
    tface.heading("col7",text="Mes",anchor=CENTER)
    tface.heading("col8",text="Año",anchor=CENTER)
    tface.heading("col9",text="Hora",anchor=CENTER)
    
    
    note.add(facn,text='Nueva \n factura')
    note.add(facm,text='Facturas\n del mes')
    note.add(facp,text='Facturas\n Pasadas')
        
def alma():
    
    def regi():
        q=0
        i=2
        w=0
        m=mm.get()
        an=aa.get()
        cadu=(f'{m}/{an}')
        if produ.get()=="" or canti.get()=="" or mm.get()=="MM" or aa.get()=="AA" or provedor.get()=="" or nomb.get()=="" or aped.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while q==0:   
            a=ht1[f'A{i}']
            ac=ht1[f'A{i}']
            v=a.value
            vv=ac.value
            if v==None:
                ht1[f'A{i}']=produ.get().lower()
                ht1[f'B{i}']=canti.get()
                ht1[f'C{i}']=cadu
                ht1[f'D{i}']=provedor.get().lower()
                ht1[f'E{i}']=nomb.get()
                ht1[f'F{i}']=aped.get()
                q=1
            
            elif(v==produ.get().lower() and vv==provedor.get().lower()):
                if ht1[f'B{i}'].value==None:
                    ht1[f'B{i}']=canti.get()
                else:
                    ht1[f'B{i}']=int(canti.get())+int(ht1[f'B{i}'].value)
                ht1[f'C{i}']=cadu
                q=1
            else:
                i=i+1
        if w==0:
            produ.set("")
            canti.set("")
            mm.set("MM")
            aa.set("AA")
            provedor.set("")
            nomb.set("")
            aped.set("")
            wb2.save('Almacen.xlsx')
        c=0
        i=2
        for item in lalma.get_children():
            lalma.delete(item)
        while c==0:
            Dt=(ht1[f'A{i}']).value
            Et=(ht1[f'B{i}']).value
            Ft=(ht1[f'C{i}']).value
            Gt=(ht1[f'D{i}']).value
            Ht=(ht1[f'E{i}']).value
            It=(ht1[f'F{i}']).value
            
            a=ht1[f'A{i}']
            v=a.value
            if v==None:
                c=1
            else:
                lalma.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It))
                i=i+1
    
    
    
    alm=Toplevel()
    alm.title("Almacén")
    note=ttk.Notebook(alm)
    note.pack(fill='both',expand='yes')
#---------------------------REGISTRO DE PRODUCTO-------------------------------
    rpo=Frame(note,bg="gray")
    produ=StringVar()
    provedor=StringVar()
    canti=StringVar()
    mm=StringVar()
    mm.set("MM")
    aa=StringVar()
    aa.set("AA")
    nomb=StringVar()
    aped=StringVar()
    prod=Label(rpo,font="Fuente 15", text="Producto:",bg="gray")
    prod.grid(row=1,column=0,padx=10,pady=10)
    pro=Entry(rpo,font="Fuente 15", width="30",textvariable=produ)
    pro.grid(row=1,column=1,columnspan=4)
    cant=Label(rpo, font="Fuente 15", text="Cantidad:",bg="gray")
    cant.grid(row=2,column=0,padx=10)
    can=Entry(rpo,font="Fuente 15", width="5",textvariable=canti)
    can.grid(row=2,column=1,padx=10,sticky=W)
    cad=Label(rpo, font="Fuente 15", text="Caducidad:",bg="gray")
    cad.grid(row=3,column=0,padx=10)
    cadm=OptionMenu(rpo, mm, *mes)
    cadm.config(font="Fuente 14")
    cadm.grid(row=3,column=1,padx=10,pady=10,sticky=W)
    cada=OptionMenu(rpo, aa, *ano)
    cada.config(font="Fuente 14")
    cada.grid(row=3,column=2,sticky=W)
    dist=Label(rpo, font="Fuente 15", text="Distribuidor:",bg="gray")
    dist.grid(row=4,column=0,padx=10)
    dis=Entry(rpo,font="Fuente 15", width="30",textvariable=provedor)
    dis.grid(row=4,column=1,columnspan=4)
    aut=Label(rpo, font="Fuente 15", text="Autorizado por:",bg="gray")
    aut.grid(row=5,column=0,padx=10,columnspan=4,pady=10)
    noma=Label(rpo, font="Fuente 15", text="Nombre:",bg="gray")
    noma.grid(row=6,column=0,padx=10)
    nom_a=Entry(rpo,font="Fuente 15", width="30",textvariable=nomb)
    nom_a.grid(row=6,column=1,padx=10,columnspan=4)
    apel=Label(rpo, font="Fuente 15", text="Apellido:",bg="gray")
    apel.grid(row=7,column=0,padx=10)
    apell=Entry(rpo,font="Fuente 15", width="30",textvariable=aped)
    apell.grid(row=7,column=1,padx=10,columnspan=4,pady=10)
    regp=Button(rpo,command=lambda:regi())
    regp.config(font="fuente 15",text="Registrar\nProducto", width="15",height=2)
    regp.grid(row=8,column=1,columnspan=2,padx=10,pady=10)
    lalma=ttk.Treeview(rpo,columns=("col1","col2","col3","col4","col5"))
    lalma.grid(row=1,column=7,padx=15,rowspan=8)
    lalma.column("#0",width="100")
    lalma.column("col1",width="50")
    lalma.column("col2",width="80")
    lalma.column("col3",width="100")
    lalma.column("col4",width="100")
    lalma.column("col5",width="100")
    lalma.heading("#0",text="Producto",anchor=CENTER)
    lalma.heading("col1",text="Cantidad",anchor=CENTER)
    lalma.heading("col2",text="caducidad",anchor=CENTER)
    lalma.heading("col3",text="Distribuidor",anchor=CENTER)
    lalma.heading("col4",text="Nombre",anchor=CENTER)
    lalma.heading("col5",text="Apellido",anchor=CENTER)
    c=0
    i=2
    while c==0:
        Dt=(ht1[f'A{i}']).value
        Et=(ht1[f'B{i}']).value
        Ft=(ht1[f'C{i}']).value
        Gt=(ht1[f'D{i}']).value
        Ht=(ht1[f'E{i}']).value
        It=(ht1[f'F{i}']).value
        a=ht1[f'A{i}']
        v=a.value
        if v==None:
            c=1
        else:
            lalma.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It))
            i=i+1
    
#-----------------------------Uso de producto----------------------------------
    rpou=Frame(note,bg="gray")
    def regus():
        q=0
        i=2
        w=0
        p=0
        tt=0
        dd=diaf.get()
        m=mesf.get()
        an=anof.get()
        fechu=(f'{dd}/{m}/{an}')
        if produc.get()=="" or cantidad.get()=="" or mesf.get()=="MM" or anof.get()=="AA" or diaf.get()=="DD" or nomru.get()=="" or aperu.get()=="" or contu.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while q==0:   
            a=ht1[f'A{i}']
            v=a.value
            if v==produc.get().lower():
                t=ht1[f'B{i}']
                to=int(t.value)
                if to=="0":
                    messagebox.showinfo("ADVERTENCIA","Sin Productos Existente")
                    q=1
                else:
                    ht2[f'A{i}']=produc.get().lower()
                    ht2[f'B{i}']=cantidad.get()
                    ht2[f'C{i}']=nomru.get()
                    ht2[f'D{i}']=aperu.get()
                    ht2[f'E{i}']=contu.get()
                    ht2[f'F{i}']=fechu
                    cada=int((ht1[f'B{i}']).value)
                    use=int(cantidad.get())
                    ht1[f'B{i}']=cada-use
                    q=1
            elif v==None:
                q=1
                p=1
            else:
                i=i+1
        if p==1:
            messagebox.showinfo("ADVERTENCIA","No se tiene registrado este producto")
        wb2.save('Almacen.xlsx')
        nomru.set("")
        aperu.set("")
        contu.set("")
        diaf.set("")
        diaf.set("DD")
        mesf.set("MM")
        anof.set("AA")
        produc.set("")
        cantidad.set("")
    
    nomru=StringVar()
    aperu=StringVar()
    contu=StringVar()
    diaf=StringVar()
    diaf.set("DD")
    mesf=StringVar()
    mesf.set("MM")
    anof=StringVar()
    anof.set("AA")
    cantidad=StringVar()
    produc=StringVar()
    nomu=Label(rpou, font="Fuente 15", text="Nombre:",bg="gray")
    nomu.grid(row=1,column=0,padx=10)
    nom_u=Entry(rpou,font="Fuente 15", width="30",textvariable=nomru)
    nom_u.grid(row=1,column=1,padx=10,columnspan=4,pady=10)
    apeu=Label(rpou, font="Fuente 15", text="Apellido:",bg="gray")
    apeu.grid(row=2,column=0,padx=10)
    apelu=Entry(rpou,font="Fuente 15", width="30",textvariable=aperu)
    apelu.grid(row=2,column=1,padx=10,columnspan=4)
    cont=Label(rpou, font="Fuente 15", text="contacto:",bg="gray")
    cont.grid(row=3,column=0,padx=10)
    conta=Entry(rpou,font="Fuente 15", width="30",textvariable=contu)
    conta.grid(row=3,column=1,padx=10,columnspan=4,pady=10)
    fecha=Label(rpou, font="Fuente 15", text="Fecha:",bg="gray")
    fecha.grid(row=4,column=0,padx=10,pady=20)
    fdia=OptionMenu(rpou,diaf,*dia)
    fdia.config(font="Fuente 14")
    fdia.grid(row=4,column=1)
    fmes=OptionMenu(rpou,mesf,*mes)
    fmes.config(font="Fuente 14")
    fmes.grid(row=4,column=2)
    fano=OptionMenu(rpou,anof,*ano)
    fano.config(font="Fuente 14")
    fano.grid(row=4,column=3)
    totl=Label(rpou, font="Fuente 15", text="Cantidad:",bg="gray")
    totl.grid(row=5,column=0,padx=10)
    tol=Entry(rpou,font="Fuente 15", width="8",textvariable=cantidad)
    tol.grid(row=5,column=1)
    prod=Label(rpou, font="Fuente 15", text="Producto:",bg="gray")
    prod.grid(row=6,column=0,padx=10,pady=10)
    pro=Entry(rpou,font="Fuente 15", width="30",textvariable=produc)
    pro.grid(row=6,column=1,columnspan=4)
    regu=Button(rpou,command=lambda:regus())
    regu.config(font="fuente 15",text="Registrar\nUso", width="15",height=2)
    regu.grid(row=7,column=1,columnspan=2,padx=10,pady=10)
    
    
    
    
#-----------------------------------Pedido-------------------------------------
    ped=Frame(note,bg="gray")
    
    def regpe():
        q=0
        i=2
        w=0
        if prov.get()=="" or producto.get()=="" or datd.get()=="DD" or datm.get()=="MM" or data.get()=="AA" or nump.get()=="" or pza.get()=="" or distr.get()=="" or cost.get()=="" or autori.get()=="":
            q=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el Pedido")
            w=1
        while q==0:   
            a=ht3[f'A{i}']
            v=a.value
            if v==nump.get():
                messagebox.showinfo("ADVERTENCIA","Numero de pedido ya asociado")
                q=1
            elif v==None:
                DD=datd.get()
                MM=datm.get()
                AA=data.get()
                datfi=f'{DD}/{MM}/{AA}'
                ht3[f'A{i}']=nump.get()
                ht3[f'B{i}']=prov.get()
                ht3[f'C{i}']=producto.get()
                ht3[f'D{i}']=pza.get()
                ht3[f'E{i}']=distr.get()
                ht3[f'F{i}']=cost.get()
                ht3[f'G{i}']=datfi
                ht3[f'H{i}']=hora.get()
                ht3[f'I{i}']=autori.get()
                ht3[f'J{i}']="Pendiente"
                q=1
            else:
                i=i+1
        if w==0:
            nump.get()
            prov.get()
            produc.get()
            pza.get()
            distr.get()
            cost.get()
            hora.get()
            autori.get()
            wb2.save('Almacen.xlsx')
        q=0
        for item in lalma.get_children():
            lped.delete(item)
        while q==0:
            Dt=(ht3[f'A{i}']).value
            Et=(ht3[f'B{i}']).value
            Ft=(ht3[f'C{i}']).value
            Gt=(ht3[f'D{i}']).value
            Ht=(ht3[f'E{i}']).value
            It=(ht3[f'F{i}']).value
            Jt=(ht3[f'G{i}']).value
            Kt=(ht3[f'H{i}']).value
            Lt=(ht3[f'I{i}']).value
            Mt=(ht3[f'J{i}']).value
            
            a=ht1[f'A{i}']
            v=a.value
            if v==None:
                q=1
            else:
                lped.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It,Jt,Kt,Lt,Mt))
                i=i+1
    
    def recibido():
        selected = lped.focus()
        temp = lped.item(selected, 'values')
        lped.item(selected, values=(temp[0], temp[1], temp[2],temp[3],temp[4], temp[5], temp[6],temp[7],"Recibido"))
        valin=lped.item(selected, 'text')
        q=0
        i=2
        while q==0:   
            a=ht3[f'A{i}']
            v=a.value
            if v==valin:
                ht3[f'J{i}']="Recibido"
                q=1
            else:
                i=i+1
            wb2.save('Almacen.xlsx')
    prov=StringVar()
    autori=StringVar()
    producto=StringVar()
    datd=StringVar()
    datm=StringVar()
    data=StringVar()
    datd.set("DD")
    datm.set("MM")
    data.set("AA")
    hora=StringVar()
    hora.set(horas[0])
    nump=StringVar()
    pza=StringVar()
    distr=StringVar()
    cost=StringVar()
    prod=Label(ped,font="Fuente 15", text="Producto",bg="gray")
    prod.grid(row=0,column=0,padx=10,pady=10)
    prou=Entry(ped,font="Fuente 15", width="30",textvariable=producto)
    prou.grid(row=0,column=1,columnspan=5,padx=10)
    prove=Label(ped,font="Fuente 15", text="Provedor:",bg="gray")
    prove.grid(row=1,column=0,padx=10)
    proved=Entry(ped,font="Fuente 15", width="30",textvariable=prov)
    proved.grid(row=1,column=1,columnspan=5,padx=10)
    fecha=Label(ped, font="Fuente 15", text="Fecha:",bg="gray")
    fecha.grid(row=2,column=0,padx=10)
    fdia=OptionMenu(ped,datd,*dia)
    fdia.config(font="Fuente 14")
    fdia.grid(row=2,column=1,pady=10)
    fmes=OptionMenu(ped,datm,*mes)
    fmes.config(font="Fuente 14")
    fmes.grid(row=2,column=2)
    fano=OptionMenu(ped,data,*ano)
    fano.config(font="Fuente 14")
    fano.grid(row=2,column=3)
    h=Label(ped, font="Fuente 15", text="Hora:",bg="gray")
    h.grid(row=3,column=0,padx=10,pady=10)
    hrs=OptionMenu(ped,hora,*horas)
    hrs.config(font="Fuente 14")
    hrs.grid(row=3,column=1,pady=10,sticky=W,padx=10)
    nped=Label(ped,font="Fuente 15", text="N° Pedido",bg="gray")
    nped.grid(row=4,column=0,padx=10)
    n_ped=Entry(ped,font="Fuente 15", width="30",textvariable=nump)
    n_ped.grid(row=4,column=1,columnspan=5,padx=10)
    canp=Label(ped, font="Fuente 15", text="Cantidad:",bg="gray")
    canp.grid(row=5,column=0,padx=10,pady=10)
    can_p=Entry(ped,font="Fuente 15", width="5",textvariable=pza)
    can_p.grid(row=5,column=1,padx=10,sticky=W,pady=10)
    distri=Label(ped,font="Fuente 15", text="Distribuidor:",bg="gray")
    distri.grid(row=6,column=0,padx=10)
    distrib=Entry(ped,font="Fuente 15", width="30",textvariable=distr)
    distrib.grid(row=6,column=1,columnspan=5,padx=10)
    cosp=Label(ped, font="Fuente 15", text="Costo:",bg="gray")
    cosp.grid(row=7,column=0,padx=10,pady=10)
    cos_p=Entry(ped,font="Fuente 15", width="10",textvariable=cost)
    cos_p.grid(row=7,column=1,padx=10,sticky=W,pady=10)
    auto=Label(ped, font="Fuente 15", text="Atorizacion:",bg="gray")
    auto.grid(row=8,column=0,padx=10)
    aut_o=Entry(ped,font="Fuente 15", width="30",textvariable=autori)
    aut_o.grid(row=8,column=1,padx=10,columnspan=5)
    regp=Button(ped,command=lambda:regpe())
    regp.config(font="fuente 15",text="Registrar\nPedido", width="15",height=2)
    regp.grid(row=9,column=1,columnspan=2,padx=10,pady=10)
    reci=Button(ped,command=lambda:recibido())
    reci.config(font="fuente 15",text="Recibido", width="15")
    reci.grid(row=8,column=7,padx=10,pady=10)
    lped=ttk.Treeview(ped,columns=("col1","col2","col3","col4","col5","col6","col7","col8","col9"))
    lped.grid(row=1,column=7,padx=15,rowspan=8)
    lped.column("#0",width="80")
    lped.column("col1",width="100")
    lped.column("col2",width="100")
    lped.column("col3",width="80")
    lped.column("col4",width="100")
    lped.column("col5",width="80")
    lped.column("col6",width="80")
    lped.column("col7",width="80")
    lped.column("col8",width="100")
    lped.column("col9",width="100")
    lped.heading("#0",text="N° Pedido",anchor=CENTER)
    lped.heading("col1",text="Provedor",anchor=CENTER)
    lped.heading("col2",text="Producto",anchor=CENTER)
    lped.heading("col3",text="Cantidad",anchor=CENTER)
    lped.heading("col4",text="Distribuidor",anchor=CENTER)
    lped.heading("col5",text="Costo",anchor=CENTER)
    lped.heading("col6",text="Fecha",anchor=CENTER)
    lped.heading("col7",text="Hora",anchor=CENTER)
    lped.heading("col8",text="Nombre",anchor=CENTER)
    lped.heading("col9",text="Estatus",anchor=CENTER)
    q=0
    i=2
    while q==0:
        Dt=(ht3[f'A{i}']).value
        Et=(ht3[f'B{i}']).value
        Ft=(ht3[f'C{i}']).value
        Gt=(ht3[f'D{i}']).value
        Ht=(ht3[f'E{i}']).value
        It=(ht3[f'F{i}']).value
        Jt=(ht3[f'G{i}']).value
        Kt=(ht3[f'H{i}']).value
        Lt=(ht3[f'I{i}']).value
        Mt=(ht3[f'J{i}']).value
        
        a=ht3[f'A{i}']
        v=a.value
        if v==None:
            q=1
        else:
            lped.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It,Jt,Kt,Lt,Mt))
            i=i+1
    
    note.add(rpo,text='Registro de\n producto')
    note.add(rpou,text='Registro de\n producto en uso')
    note.add(ped,text='Pedido de\n Producto')
    
def faca():
    faca=Toplevel()
    faca.title("Finanzas del almacen")
    note= ttk.Notebook(faca)
    note.pack(fill='both',expand='yes')
    facn=Frame(note,bg="gray")
    def agref():
        c=0
        i=2
        w=0
        fed=daye.get()
        fem=monthe.get()
        fea=yeae.get()
        fpd=dayp.get()
        fpm=monthp.get()
        fpa=yeap.get()
        cam=cm.get()
        caa=ca.get()
        pedif=(f"{fed}/{fem}/{fea}")
        entref=(f"{fpd}/{fpm}/{fpa}")
        caduf=(f"{cam}/{caa}")
        if numefa.get()=="" or nombfa.get()=="" or apelfa.get()=="" or montfa.get()=="" or areafa.get()=="" or encafa.get()=="" :
            c=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while c==0:   
            a=ht4[f'A{i}']
            v=a.value
            if v==None:
                ht4[f'A{i}']=numefa.get()
                ht4[f'B{i}']=nombfa.get()
                ht4[f'C{i}']=entref
                ht4[f'D{i}']=pedif
                ht4[f'E{i}']=apelfa.get()
                ht4[f'F{i}']=areafa.get()
                ht4[f'G{i}']=caduf
                ht4[f'H{i}']=encafa.get()
                ht4[f'I{i}']=montfa.get()
                ht4[f'J{i}']=monthe.get()
                c=1
            else:
                i=i+1
        if w==0:
            numefa.set("")
            nombfa.set("")
            apelfa.set("")
            areafa.set("")
            montfa.set("")
            encafa.set("")
            daye.set("DD")
            monthe.set("MM")
            yeae.set("AA")
            dayp.set("DD")
            monthp.set("MM")
            yeap.set("AA")
            cm.set("MM")
            ca.set("AA")
            
        wb2.save('Almacen.xlsx')
        
    numefa=StringVar()
    nombfa=StringVar()
    apelfa=StringVar()
    areafa=StringVar()
    montfa=StringVar()
    encafa=StringVar()
    daye=StringVar()
    monthe=StringVar()
    yeae=StringVar()
    daye.set("DD")
    monthe.set("MM")
    yeae.set("AA")
    dayp=StringVar()
    monthp=StringVar()
    yeap=StringVar()
    dayp.set("DD")
    monthp.set("MM")
    yeap.set("AA")
    cm=StringVar()
    ca=StringVar()
    cm.set("MM")
    ca.set("AA")
    num=Label(facn,font="Fuente 15", text="N° Pedido:",bg="gray")
    num.grid(row=1,column=0,padx=10,pady=15)
    nume=Entry(facn,font="Fuente 15", width="10",textvariable=numefa)
    nume.grid(row=1,column=1,padx=10,sticky=W)
    nomb=Label(facn,font="Fuente 15", text="Provedor:",bg="gray")
    nomb.grid(row=2,column=0,padx=10)
    nom=Entry(facn,font="Fuente 15", width="30",textvariable=nombfa)
    nom.grid(row=2,column=1,columnspan=4,padx=10)
    fechap=Label(facn, font="Fuente 15", text="Fecha de\nPedido:",bg="gray")
    fechap.grid(row=3,column=0,padx=10,pady=10)
    fdiap=OptionMenu(facn,dayp,*dia)
    fdiap.config(font="Fuente 14")
    fdiap.grid(row=3,column=1,pady=10)
    fmesp=OptionMenu(facn,monthp,*mes)
    fmesp.config(font="Fuente 14")
    fmesp.grid(row=3,column=2,pady=10,sticky=W)
    fanop=OptionMenu(facn,yeap,*ano)
    fanop.config(font="Fuente 14")
    fanop.grid(row=3,column=3)
    fechae=Label(facn, font="Fuente 15", text="Fecha de\nEntrega:",bg="gray")
    fechae.grid(row=4,column=0,padx=10)
    fdiae=OptionMenu(facn,daye,*dia)
    fdiae.config(font="Fuente 14")
    fdiae.grid(row=4,column=1)
    fmese=OptionMenu(facn,monthe,*mes)
    fmese.config(font="Fuente 14")
    fmese.grid(row=4,column=2,sticky=W)
    fanoe=OptionMenu(facn,yeae,*ano)
    fanoe.config(font="Fuente 14")
    fanoe.grid(row=4,column=3)
    apel=Label(facn,font="Fuente 15", text="Producto:",bg="gray")
    apel.grid(row=5,column=0,padx=10,pady=10)
    ape=Entry(facn,font="Fuente 15", width="30",textvariable=apelfa)
    ape.grid(row=5,column=1,columnspan=4,padx=10)
    cant=Label(facn,font="Fuente 15",text="Cantidad:",bg="gray")
    cant.grid(row=6,column=0,padx=10)
    can=Entry(facn,font="Fuente 15", width="10",textvariable=areafa)
    can.grid(row=6,column=1,sticky=W,pady=10,padx=10)
    cad=Label(facn, font="Fuente 15", text="Caducidad:",bg="gray")
    cad.grid(row=7,column=0,padx=10)
    cadm=OptionMenu(facn, cm, *mes)
    cadm.config(font="Fuente 14")
    cadm.grid(row=7,column=1,padx=10,sticky=W,pady=10)
    cada=OptionMenu(facn, ca, *ano)
    cada.config(font="Fuente 14")
    cada.grid(row=7,column=2)
    area=Label(facn, font="Fuente 15", text="Distribuidor:",bg="gray")
    area.grid(row=8,column=0,padx=10,pady=10)
    are=Entry(facn,font="Fuente 15", width="30",textvariable=encafa)
    are.grid(row=8,column=1,columnspan=4,pady=10)
    mont=Label(facn,font="Fuente 15",text="Monto:",bg="gray")
    mont.grid(row=9,column=0,padx=10)
    mon=Entry(facn,font="Fuente 15", width="10",textvariable=montfa)
    mon.grid(row=9,column=1,sticky=W,padx=10,pady=10)
    ingre=Button(facn,command=lambda:agref())
    ingre.config(font="fuente 15",text="Guardar Factura", width="15")
    ingre.grid(row=5,column=6,padx=15)
    
    facm=Frame(note,bg="gray")
    def busc():
        c=0
        i=2
        w=0
        if femo.get()=="MM" or femo.get()=="XX":
            c=1
            messagebox.showinfo("ADVERTENCIA","Datos faltantes para competar el registro")
            w=1
        while c==0:
            D=(ht4[f'A{i}']).value
            E=(ht4[f'B{i}']).value
            F=(ht4[f'C{i}']).value
            G=(ht4[f'D{i}']).value
            H=(ht4[f'E{i}']).value
            I=(ht4[f'F{i}']).value
            J=(ht4[f'G{i}']).value
            K=(ht4[f'H{i}']).value
            L=(ht4[f'I{i}']).value
            
            a=ht4[f'J{i}']
            v=a.value
            if v==femo.get():
                lface.insert("",END,text=D,values=(E,F,G,H,I,J,K,L))
                i=i+1
            elif v==None:
                c=1
            else:
                i=i+1
        if w==0:
            femo.set("MM")
            
    femo=StringVar()
    femo.set("MM")
    fme=Label(facm,font="Fuente 15", text="Mes:",bg="gray")
    fme.grid(row=1,column=0,padx=10,pady=15)
    fmes=OptionMenu(facm,femo,*mes)
    fmes.config(font="Fuente 14")
    fmes.grid(row=1,column=1,pady=10,sticky=W)
    bus=Button(facm,command=lambda:busc())
    bus.config(font="fuente 15",text="Buscar Facturas", width="15")
    bus.grid(row=1,column=2,padx=15)
    lface=ttk.Treeview(facm,columns=("col1","col2","col3","col4","col5","col6","col7","col8"))
    lface.grid(row=2,column=0,columnspan=4,padx=15)
    lface.column("#0",width="50")
    lface.column("col1",width="100")
    lface.column("col2",width="80")
    lface.column("col3",width="80")
    lface.column("col4",width="100")
    lface.column("col5",width="50")
    lface.column("col6",width="50")
    lface.column("col7",width="100")
    lface.column("col8",width="50")
    lface.heading("#0",text="N° ped",anchor=CENTER)
    lface.heading("col1",text="Provedor",anchor=CENTER)
    lface.heading("col2",text="Fecha pedido",anchor=CENTER)
    lface.heading("col3",text="Fecha entrega",anchor=CENTER)
    lface.heading("col4",text="producto",anchor=CENTER)
    lface.heading("col5",text="Cantidad",anchor=CENTER)
    lface.heading("col6",text="Caducidad",anchor=CENTER)
    lface.heading("col7",text="Distribuidor",anchor=CENTER)
    lface.heading("col8",text="Monto",anchor=CENTER)
    
    
    facp=Frame(note,bg="gray")
    def carf():
        c=0
        i=2
        while c==0:
            Dt=(ht4[f'A{i}']).value
            Et=(ht4[f'B{i}']).value
            Ft=(ht4[f'C{i}']).value
            Gt=(ht4[f'D{i}']).value
            Ht=(ht4[f'E{i}']).value
            It=(ht4[f'F{i}']).value
            Jt=(ht4[f'G{i}']).value
            Kt=(ht4[f'H{i}']).value
            Lt=(ht4[f'I{i}']).value
            
            a=ht4[f'A{i}']
            v=a.value
            if v==None:
                c=1
            else:
                tface.insert("",END,text=Dt,values=(Et,Ft,Gt,Ht,It,Jt,Kt,Lt))
                i=i+1
    car=Button(facp,command=lambda:carf())
    car.config(font="fuente 15",text="Buscar Facturas", width="15")
    car.grid(row=1,column=1,padx=15,pady=15)
    tface=ttk.Treeview(facp,columns=("col1","col2","col3","col4","col5","col6","col7","col8"))
    tface.grid(row=2,column=0,columnspan=4,padx=15)
    tface.column("#0",width="50")
    tface.column("col1",width="100")
    tface.column("col2",width="80")
    tface.column("col3",width="80")
    tface.column("col4",width="100")
    tface.column("col5",width="50")
    tface.column("col6",width="50")
    tface.column("col7",width="100")
    tface.column("col8",width="50")
    tface.heading("#0",text="N° ped",anchor=CENTER)
    tface.heading("col1",text="Provedor",anchor=CENTER)
    tface.heading("col2",text="Fecha pedido",anchor=CENTER)
    tface.heading("col3",text="Fecha entrega",anchor=CENTER)
    tface.heading("col4",text="producto",anchor=CENTER)
    tface.heading("col5",text="Cantidad",anchor=CENTER)
    tface.heading("col6",text="Caducidad",anchor=CENTER)
    tface.heading("col7",text="Distribuidor",anchor=CENTER)
    tface.heading("col8",text="Monto",anchor=CENTER)
    

    note.add(facn,text='Nueva \n factura')
    note.add(facm,text='Facturas\n del mes')
    note.add(facp,text='Facturas\n Pasadas')
    

inicio=Frame()
inicio.pack()
inicio.config(bg="gray")
inicio.config()
intro=Label(inicio,font="fuente 20",text="Bienvenidos al sistema de administración de personal:\nSeleccione a donde quiere ir",bg="gray")
intro.grid(row=1,column=0,padx=20,pady=15,columnspan=2)
rh=Button(inicio,command=lambda:rec())
rh.config(font="fuente 14",text="Registro de\nAsistencia", width="18",height=2)
rh.grid(row=2,column=0)
almacen=Button(inicio,command=lambda:alma())
almacen.config(font="fuente 14",text="Almacén", width="18",height=2)
almacen.grid(row=2,column=1)
fac_emp=Button(inicio,command=lambda:face())
fac_emp.config(font="fuente 14",text="Pagos a\nEmpleados/Facturas", width="18",height=2)
fac_emp.grid(row=3,column=0,pady=20)
fac_alm=Button(inicio,command=lambda:faca())
fac_alm.config(font="fuente 14",text="Facturas de\n Almacén", width="18",height=2)
fac_alm.grid(row=3,column=1,pady=20)
home.mainloop()